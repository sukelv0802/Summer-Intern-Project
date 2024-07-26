import sys
import time
import serial
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.chart import LineChart, Reference
from PyQt5.QtWidgets import (QApplication, QMainWindow, QTreeWidget, QTreeWidgetItem, 
                             QPushButton, QVBoxLayout, QHBoxLayout, QWidget, QLabel, 
                             QLineEdit, QMessageBox, QHeaderView, QSplitter, QCheckBox,
                             QStyleFactory, QComboBox, QFileDialog, QDateTimeEdit, 
                             QStatusBar, QListWidget, QTabWidget)
from PyQt5.QtCore import QTimer, Qt, QSettings, QDateTime
from PyQt5.QtGui import QFont, QColor
import pyqtgraph as pg


"""
Main application window for the Serial Data Logger.

This class sets up the GUI, manages serial communication, and handles data logging and visualization.
"""
class MainWindow(QMainWindow):
    
    #  Initialize the main window and set up the user interface.
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Serial Data Logger")
        self.setGeometry(100, 100, 1000, 600)
        self.cycle_count = 0
        self.data_point_count = 0

        self.setup_settings()
        self.setup_ui()
        self.setup_serial_connection()
        self.setup_variables()

    # Load and apply application settings
    def setup_settings(self):
        self.settings = QSettings("Test", "SerialDataLogger")
        self.move(self.settings.value("pos", self.pos()))
        self.resize(self.settings.value("size", self.size()))

    # Set up the user interface components
    def setup_ui(self):
        main_widget = QWidget()
        main_layout = QHBoxLayout(main_widget)
        splitter = QSplitter(Qt.Horizontal)
        
        options_bar = self.create_options_bar()
        self.tab_widget = self.create_tab_widget()
        
        splitter.addWidget(options_bar)
        splitter.addWidget(self.tab_widget)
        splitter.setStretchFactor(1, 1)
        
        main_layout.addWidget(splitter)
        self.setCentralWidget(main_widget)
        
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)

    # Create and return the options sidebar
    def create_options_bar(self):
        options_bar = QWidget()
        options_layout = QVBoxLayout(options_bar)
        options_layout.setAlignment(Qt.AlignTop)
        
        self.add_threshold_widgets(options_layout)
        self.add_cycle_period_widgets(options_layout)
        self.add_com_port_widget(options_layout)
        self.add_mux_selection_widget(options_layout)
        self.add_channel_selection_widget(options_layout)
        
        options_layout.addStretch()
        return options_bar

    # Add threshold-related widgets to the given layout
    def add_threshold_widgets(self, layout):
        threshold_label = QLabel("Threshold")
        threshold_label.setFont(QFont("Arial", 12, QFont.Bold))
        self.threshold_entry = QLineEdit()
        self.threshold_entry.setPlaceholderText("Enter threshold value")
        confirm_button = QPushButton("Set Threshold")
        confirm_button.clicked.connect(self.set_threshold)
        
        layout.addWidget(threshold_label)
        layout.addWidget(self.threshold_entry)
        layout.addWidget(confirm_button)

    # Add Cycle period related widgets to the given layout
    def add_cycle_period_widgets(self, layout):
        cycle_period_label = QLabel("Cycle Period(s)")
        cycle_period_label.setFont(QFont("Arial", 12, QFont.Bold))
        self.cycle_period_entry = QLineEdit()
        self.cycle_period_entry.setPlaceholderText("Enter cycle period in seconds")
        set_cycle_period_button = QPushButton("Set Cycle Period")
        set_cycle_period_button.clicked.connect(self.set_cycle_period)
        
        layout.addWidget(cycle_period_label)
        layout.addWidget(self.cycle_period_entry)
        layout.addWidget(set_cycle_period_button)

    #  Add COM port selection widget to the given layout
    def add_com_port_widget(self, layout):
        com_label = QLabel("COM Port")
        com_label.setFont(QFont("Arial", 12, QFont.Bold))
        self.com_combo = QComboBox()
        self.com_combo.addItems([f"COM{i}" for i in range(1, 21)])
        self.com_combo.setCurrentText(self.settings.value("com_port", "COM10"))
        
        layout.addWidget(com_label)
        layout.addWidget(self.com_combo)

    # Add multiplexer selection widget to the given layout
    def add_mux_selection_widget(self, layout):
        mux_label = QLabel("Select Mux")
        mux_label.setFont(QFont("Arial", 12, QFont.Bold))
        self.mux_combo = QComboBox()
        self.mux_combo.addItems([f"Mux {i}" for i in range(1, 9)])
        self.mux_combo.currentIndexChanged.connect(self.update_plot)
        
        layout.addWidget(mux_label)
        layout.addWidget(self.mux_combo)

    # Add channel selection widget to the given layout
    def add_channel_selection_widget(self, layout):
        channel_label = QLabel("Select Channels")
        channel_label.setFont(QFont("Arial", 12, QFont.Bold))
        self.channel_list = QListWidget()
        self.channel_list.setSelectionMode(QListWidget.MultiSelection)
        for i in range(1, 33):
            self.channel_list.addItem(f"Channel {i}")
        self.channel_list.itemSelectionChanged.connect(self.update_plot)
        
        layout.addWidget(channel_label)
        layout.addWidget(self.channel_list)

    # Create and return the tab widget containing data and plot tabs
    def create_tab_widget(self):
        tab_widget = QTabWidget()
        
        data_tab = self.create_data_tab()
        plot_tab = self.create_plot_tab()
        
        tab_widget.addTab(data_tab, "Data")
        tab_widget.addTab(plot_tab, "Plot")
        
        return tab_widget
    
    # Create and return the data tab widget
    def create_data_tab(self):
        data_widget = QWidget()
        data_layout = QVBoxLayout(data_widget)
        
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["Timestamp", "Mux", "Channel", "Temperature", "Voltage"])
        self.tree.header().setSectionResizeMode(QHeaderView.Interactive)
        self.tree.setAlternatingRowColors(True)
        self.tree.setSortingEnabled(True)
        
        filter_layout = QHBoxLayout()
        self.filter_checkbox = QCheckBox("Show only above threshold")
        self.filter_checkbox.stateChanged.connect(self.apply_filter)
        filter_layout.addWidget(self.filter_checkbox)        
        
        button_layout = QHBoxLayout()
        self.start_button = QPushButton("Start")
        self.resume_button = QPushButton("Resume")
        self.stop_button = QPushButton("Stop")
        export_button = QPushButton("Export to Excel")
        clear_button = QPushButton("Clear Data")
        
        self.start_button.clicked.connect(self.start_update)
        self.resume_button.clicked.connect(self.resume_update)
        self.stop_button.clicked.connect(self.stop_update)
        export_button.clicked.connect(self.export_to_excel)
        clear_button.clicked.connect(self.clear_data)
        
        button_layout.addWidget(self.start_button)
        button_layout.addWidget(self.resume_button)
        button_layout.addWidget(self.stop_button)
        button_layout.addWidget(export_button)
        button_layout.addWidget(clear_button)
        
        data_layout.addWidget(self.tree)
        data_layout.addLayout(button_layout)
        data_layout.addLayout(filter_layout)
        
        return data_widget

    # Create and return the plot tab widget
    def create_plot_tab(self):
        plot_widget = QWidget()
        plot_layout = QVBoxLayout(plot_widget)
        
        self.plot = pg.PlotWidget()
        self.plot.setBackground('w')
        self.plot.setTitle("Real-time Voltage Plot")
        self.plot.setLabel('left', "Voltage")
        self.plot.setLabel('bottom', "Data Point Count")
        self.plot.addLegend()
        
        self.plot.setMouseEnabled(x=True, y=True)
        self.plot.enableAutoRange()
        
        plot_layout.addWidget(self.plot)
        
        return plot_widget

    # Setup the serial connection parameters
    def setup_serial_connection(self):
        self.baudrate = 115200
        self.serialConnection = None

    # Initialize various variables used in the application
    def setup_variables(self):
        self.threshold_value = None
        self.update_flag = False
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_text)
        self.start_time = None
        self.channel_period_value = 50
        self.cycle_period_value = 60
        self.plot_data = {}
        self.current_mux = 1

    # Handles the window close event
    def closeEvent(self, event):
        self.settings.setValue("pos", self.pos())
        self.settings.setValue("size", self.size())
        self.settings.setValue("com_port", self.com_combo.currentText())
        super().closeEvent(event)

    # Sets the threshold value for data filtering
    def set_threshold(self):
        try:
            self.threshold_value = float(self.threshold_entry.text())
            QMessageBox.information(self, "Success", f"Threshold set to {self.threshold_value}")
            self.apply_filter()  # Reapply filter after changing threshold
        except ValueError:
            QMessageBox.warning(self, "Invalid Input", "Please enter a valid value for the threshold.")

    # Set cycle period for data collectiong (cycle period includes the time to switch through all 256 channels as well as the waiting time afterwards)
    def set_cycle_period(self):
        try:
            self.cycle_period_value = int(self.cycle_period_entry.text())
            QMessageBox.information(self, "Success", f"Cycle period set to {self.cycle_period_value} seconds")
        except ValueError:
            QMessageBox.warning(self, "Invalid Input", "Please enter a valid value for the cycle period.")

    # Starts the data collection process
    def start_update(self):
        if not self.update_flag:
            try:
                com_port = self.com_combo.currentText()
                self.serialConnection = serial.Serial(com_port, self.baudrate)
                self.serialConnection.flush()
                self.serialConnection.write("START\r".encode())
                self.update_flag = True
                self.timer.start(self.channel_period_value) # CHANGE IF NEEDS TO BE FASTER/SLOWER, no faster than freq in main
                self.start_button.setEnabled(False)
                self.resume_button.setEnabled(False)
                self.stop_button.setEnabled(True)
                self.serialConnection.reset_input_buffer()
                self.start_time = time.time()
                self.statusBar.showMessage(f"Connected to {com_port} at {self.baudrate} baud.")
            except serial.SerialException as e:
                QMessageBox.critical(self, "Error", f"Failed to open {com_port}: {str(e)}")
                print(f"Failed to open {com_port}: {str:e}") # Debug message

    # Resumes the data collection process
    def resume_update(self):
        if not self.update_flag:
            try:
                com_port = self.com_combo.currentText()
                self.serialConnection = serial.Serial(com_port, self.baudrate)
                self.serialConnection.flush()
                self.serialConnection.write("RESUME\r".encode())
                self.update_flag = True
                self.timer.start(self.channel_period_value) # CHANGE IF NEEDS TO BE FASTER/SLOWER, no faster than freq in main
                self.start_button.setEnabled(False)
                self.resume_button.setEnabled(False)
                self.stop_button.setEnabled(True)
                self.serialConnection.reset_input_buffer()
                self.statusBar.showMessage("Connection resumed")
            except serial.SerialException as e:
                QMessageBox.critical(self, "Error", f"Failed to open {com_port}: {str(e)}")
                print(f"Failed to open {com_port}: {str:e}") # Debug message

    # Stop the data collection process
    def stop_update(self):
        self.update_flag = False
        self.timer.stop()
        if self.serialConnection:
            # Send PAUSE command
            self.serialConnection.flush()
            self.serialConnection.write("PAUSE\r".encode())
            print("Pause command sent")
            try:
                # Wait for a response with a timeout
                self.serialConnection.timeout = 0.1  # Set a timeout for reading
                response = ''
                while True:
                    response = self.serialConnection.readline().strip()
                    print(f"Response received: '{response.decode()}")
                    if 'Pause confirmed' in response.decode():
                        break
                    elif 'Mux:' in response.decode() and 'Channel:' in response.decode() and 'Temperature:' in response.decode() and 'Voltage:' in response.decode():
                        self.process_data(response.decode())
            except Exception as e:
                print(f"Error receiving confirmation: {e}")
            self.serialConnection.close()
            self.serialConnection = None
        self.start_button.setEnabled(True)
        self.resume_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.statusBar.showMessage("Connection paused")

    # Update the data display with new readings
    def update_text(self):
        if self.update_flag and self.serialConnection:
            try:
                data = self.serialConnection.readline()
                print(f"Data received: {data}") # Debug message
                if data == b"EOF":
                    self.stop_update()
                else:
                    value = data.decode('utf-8').strip()
                    print(f"Decoded value: {value}") # Debug message
                    timestamp = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
                    try:
                        if "Mux:" in value and "Channel:" in value and "Temperature:" in value and "Voltage:" in value:
                            self.process_data(value)
                    except ValueError:
                        self.tree.addTopLevelItem(QTreeWidgetItem([timestamp, value, "", "", ""]))
                        self.apply_filter()  # Apply filter even for invalid data
            except serial.SerialException as e:
                self.stop_update()
                QMessageBox.critical(self, "Error", f"Serial communication error: {str(e)}")
    
    # Process and store the received data   
    def process_data(self, value):
        parts = value.split()
        mux = int(parts[1])
        channel = int(parts[3])
        temperature = parts[5]
        voltage = float(parts[7])

        timestamp = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        item = QTreeWidgetItem([timestamp, str(mux), str(channel), f"{temperature}°C", f"{voltage}"])
        
        if self.threshold_value is not None and voltage < self.threshold_value:
            item.setBackground(4, QColor(255, 255, 0, 100))
        
        self.tree.addTopLevelItem(item)
        self.tree.scrollToBottom()
        self.apply_filter()
        
        self.data_point_count += 1

        if mux not in self.plot_data:
            self.plot_data[mux] = {}
        if channel not in self.plot_data[mux]:
            self.plot_data[mux][channel] = {'x': [], 'y': []}

        current_time = time.time() - self.start_time
        self.plot_data[mux][channel]['x'].append(self.data_point_count)
        self.plot_data[mux][channel]['y'].append(voltage)

        if channel == 31 and mux == 8:
            self.cycle_count += 1
            self.stop_update()
            cycle_delay = self.cycle_period_value * 1000 - 256 * 2 * self.channel_period_value
            if cycle_delay > 0:
                QTimer.singleShot(cycle_delay, self.start_update)
            else:
                self.start_update()

    # Update the plot with the latest data
    def update_plot(self):
        self.plot.clear()
        selected_mux = int(self.mux_combo.currentText().split()[1])
        selected_channels = [int(item.text().split()[1]) for item in self.channel_list.selectedItems()]

        if selected_mux in self.plot_data:
            for channel in self.plot_data[selected_mux]:
                if not selected_channels or channel in selected_channels:
                    x_data = np.array(self.plot_data[selected_mux][channel]['x'], dtype=int)
                    y_data = np.array(self.plot_data[selected_mux][channel]['y'], dtype=float)
                    self.plot.plot(
                        x_data,
                        y_data,
                        pen=(channel * 20) % 256,
                        name=f'Channel {channel}'
                    )

        self.plot.addLegend()

    # Export the collected data to an Excel file
    def export_to_excel(self):
        if self.tree.topLevelItemCount() == 0:
            QMessageBox.warning(self, "No Data", "There is no data to export.")
            return

        try:
            data = []
            for i in range(self.tree.topLevelItemCount()):
                item = self.tree.topLevelItem(i)
                data.append([item.text(0), item.text(1), item.text(2), item.text(3), item.text(4)])

            filename, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
            if not filename: # If the file dialog is cancelled
                return 

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Serial Data"
            sheet['A1'] = "Timestamp"
            sheet['B1'] = "Mux"
            sheet['C1'] = "Channel"
            sheet['D1'] = "Temperature"
            sheet["E1"] = "Voltage"

            # Write data and apply conditional formatting
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            for row, (timestamp, mux_index, channel_index, temperature, voltage) in enumerate(data, start=2):
                sheet.cell(row=row, column=1, value=timestamp)
                sheet.cell(row=row, column=2, value=mux_index)
                sheet.cell(row=row, column=3, value=channel_index)
                sheet.cell(row=row, column=4, value=temperature)
                sheet.cell(row=row, column=5, value=voltage)
                
                try:
                    numeric_value = float(voltage)
                    if self.threshold_value is not None and numeric_value < self.threshold_value:
                        sheet.cell(row=row, column=5).fill = yellow_fill
                except ValueError:
                    pass 

            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(filename)
            QMessageBox.information(self, "Success", f"Data has been successfully exported to {filename}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to export data: {str(e)}")
 
    # Clear all collected data and reset the display        
    def clear_data(self):
        self.tree.clear()
        self.filter_checkbox.setChecked(False)
        self.plot.clear()
        self.plot_data = {}
        self.start_time = None
        
    # Apply the threshold filter to the displayed data
    def apply_filter(self):
        for i in range(self.tree.topLevelItemCount()):
            item = self.tree.topLevelItem(i)
            if self.filter_checkbox.isChecked():
                try:
                    voltage = float(item.text(4))
                    item.setHidden(not (self.threshold_value is not None and voltage < self.threshold_value))
                except ValueError:
                    item.setHidden(True)
            else:
                item.setHidden(False)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())